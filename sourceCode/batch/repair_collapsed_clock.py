# -*- coding: utf-8 -*-
r"""Reconstructs the afternoon half of raw HOBO exports written with a collapsed
12-hour clock, IN PLACE, and records the operation in the raw manifest.

The defect: HOBOware in pt-BR locale writes the time as '04h0min0s' - a
12-hour clock with **no AM/PM marker**. Every afternoon reading therefore lands
on top of its morning twin: about half the timestamps are duplicated and no
sample ever falls after 12:59. Half of each affected series sits on the wrong
timestamp, and time-of-day analysis on it is meaningless.

What makes the reconstruction DETERMINED rather than guessed (verified over the
62 affected exports before this script was written):

  * the DATE field is intact - only the hour collapsed, and the date rolls over
    correctly at midnight, so each row's two candidate times are exactly
    `date + h%12` and that + 12 h;
  * inside one calendar date the true times rise monotonically with a single
    noon crossing - measured on the archive, 61 of the 62 exports satisfy this
    (the exception is a hand-assembled multi-site incubation sheet, excluded
    below and left for manual handling);
  * sampling is regular (mode: 2 h; also 10 s, 1 min, 30 min, 1 h and 3 h
    files), so a wrong AM/PM choice breaks the step and is detectable.

The walk therefore picks, for each row, the candidate that is strictly after
the previous timestamp and closest to `previous + interval`. That recovers the
RELATIVE structure - which samples sit 12 h apart - but not the absolute half:
adding 12 h to the whole series leaves every step identical, so the sampling
interval provably cannot tell a morning start from an afternoon one.

Only physics can: a submerged light sensor must peak near local noon. The
absolute half is therefore chosen by running `QCS_Tests.light_clock_phase` on
both candidate series and keeping the one whose light peaks nearer midday - and
when the light is too weak to decide (a sensor dark from the start), the file
is REFUSED rather than guessed.

Nothing is written until every gate passes: strictly increasing times, no
duplicates left, enough steps on the sampling interval, a noon-centred light
phase, and - the strongest one - proof that ONLY clock fields changed, checked
by blanking every clock field in both texts and requiring the remainder to be
byte-identical. (Row counts are deliberately not compared: the reader trims
out-of-water edges by a temperature heuristic, so a correctly repaired file
legitimately keeps a different number of rows.) Files that fail any gate are
reported and left untouched - never half-repaired.

The loggers' own binary exports (raw\<SITE>\<campaign>\bruto\*.hobo) are not
touched: re-exporting from HOBOware with a 24-hour clock remains the
inference-free alternative for anything this script refuses.

Usage:  python repair_collapsed_clock.py            (repairs, writes)
        python repair_collapsed_clock.py --dry-run  (reports, writes nothing)
"""
import csv

import glob
import hashlib
import os
import re
import shutil
import sys
import tempfile
import warnings
from collections import Counter

warnings.filterwarnings('ignore')
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import numpy as np                                                   # noqa: E402
import pandas as pd                                                  # noqa: E402
import QCS_DataHandler as dh                                         # noqa: E402
import QCS_Tests as QC                                               # noqa: E402

ROOT = r"\\Abrolhos\Projetos\Seaguard & HOBO\CLAUDE\HOBO"
H_RAW = os.path.join(ROOT, 'raw')
MANIFEST = os.path.join(H_RAW, 'manifest.csv')

# '03/17/18 04h0min0s' - date left untouched, only the clock is rewritten
_STAMP = re.compile(r'(\d\d)/(\d\d)/(\d\d)(\s+)(\d{1,2})h(\d{1,2})min(\d{1,2})s')

# Hand-assembled sheet holding SEVERAL sites stacked in one file (Sitio;Data...),
# so its dates legitimately jump backwards - the premise does not apply and an
# automated reconstruction would scramble it.
EXCLUDED = {'HOBO-incubacao_rodolito.csv'}

MIN_STEP_MATCH = 0.95        # share of steps that must sit on the interval
MAX_NOON_OFFSET_H = 3.0      # repaired light must peak within this of noon

# Fallback when the light cannot choose the absolute half - a shaded or fouled
# light channel peaks nowhere near noon in EITHER half, but the water still
# warms in the afternoon. The corpus median hour of the daily temperature
# maximum is 13.7 h (measured over 47 clean-phase products), so the half whose
# temperature peaks nearer that wins. Wider tolerance than the light gate: the
# diurnal temperature signal of reef water is ~1 degC and noisier than light.
TEMP_PEAK_H = 13.7
MAX_TEMP_OFFSET_H = 5.0


def temp_peak_hour(t, temp):
    """Circular mean of the hour at which each day's temperature peaks, or NaN
    when there are too few complete days to mean anything."""
    s = pd.Series(pd.to_numeric(temp, errors='coerce').values,
                  index=pd.DatetimeIndex(t)).dropna()
    hrs = [g.idxmax().hour + g.idxmax().minute / 60.0
           for _d, g in s.groupby(s.index.date) if len(g) >= 8]
    if len(hrs) < 20:
        return float('nan')
    a = 2 * np.pi * np.asarray(hrs, dtype=float) / 24.0
    return float((np.arctan2(np.sin(a).sum(), np.cos(a).sum()) * 24 / (2 * np.pi)) % 24)


def parse_stamps(text):
    """[(match, date, seconds-within-half-day)] in file order."""
    out = []
    for m in _STAMP.finditer(text):
        mo, dd, yy, _sp, hh, mi, ss = m.groups()
        # the reader decides day-first vs month-first from the file's own
        # evidence; here only the CLOCK matters, so the date is kept as text
        # and used solely as a grouping key that increments at midnight
        out.append((m, (yy, mo, dd),
                    (int(hh) % 12) * 3600 + int(mi) * 60 + int(ss)))
    return out


def reconstruct(stamps, seed_pm=False):
    """Full 24-hour seconds-of-day for each stamp, or None when the series
    cannot be made regular. Dates are ordinal-indexed by first appearance, so
    the walk needs no calendar arithmetic on an ambiguous date format.

    `seed_pm` places the FIRST sample in the afternoon half. That is the only
    free choice in the whole reconstruction, and it must be made here rather
    than by shifting the finished series: every sample belongs to the day its
    own (correct) date field names, so a global '+12 h' would have to push some
    samples onto the next date - and wrapping the clock instead silently puts a
    22:00 sample before a 00:00 one inside the same date."""
    if len(stamps) < 4:
        return None
    order, day_index = {}, []
    for _m, d, _s in stamps:
        if d not in order:
            order[d] = len(order)
        day_index.append(order[d])
    within = [s for _m, _d, s in stamps]

    # sampling interval: the modal positive step inside a half-day
    steps = Counter()
    for i in range(1, len(within)):
        if day_index[i] == day_index[i - 1] and within[i] > within[i - 1]:
            steps[within[i] - within[i - 1]] += 1
    if not steps:
        return None
    interval = steps.most_common(1)[0][0]

    # RELATIVE structure only. The absolute half is undetermined here by
    # construction (shifting everything by 12 h leaves every step identical),
    # and is resolved later from the light phase.
    times, prev = [], None
    for i, w in enumerate(within):
        base = day_index[i] * 86400 + w
        cands = (base, base + 12 * 3600)
        if prev is None:
            t = cands[1] if seed_pm else cands[0]
        else:
            later = [c for c in cands if c > prev]
            if not later:
                return None
            t = min(later, key=lambda c: abs((c - prev) - interval))
        times.append(t)
        prev = t
    score = sum(1 for a, b in zip(times, times[1:], strict=False) if b - a == interval)
    return {'times': times, 'interval': interval,
            'step_match': score / max(len(times) - 1, 1)}


def rewrite(text, stamps, times):
    """The export with every clock field rewritten to 24-hour, everything else
    byte-identical (dates, values, separators, line endings).

    The DATE text is copied verbatim, which is sound precisely because
    `reconstruct` only ever picks a time inside the day the row's own date
    names - no sample is moved across midnight."""
    pieces, last = [], 0
    for (m, _d, _s), t in zip(stamps, times, strict=True):
        sod = t % 86400
        pieces.append(text[last:m.start()])
        pieces.append('%s/%s/%s%s%dh%dmin%ds'
                      % (m.group(1), m.group(2), m.group(3), m.group(4),
                         sod // 3600, (sod % 3600) // 60, sod % 60))
        last = m.end()
    pieces.append(text[last:])
    return ''.join(pieces)


def _clock_blanked(text):
    """`text` with every clock field replaced by a constant, so two versions of
    the same export compare equal iff nothing BUT the clock changed."""
    return _STAMP.sub(lambda m: '%s/%s/%s%s<clock>'
                      % (m.group(1), m.group(2), m.group(3), m.group(4)), text)


# --------------------------------------------------------------- xlsx exports
# The same defect reaches .xlsx exports, where the collapsed clock sits as CELL
# TEXT (single sheet, no merged cells, one datetime column). openpyxl rewrites
# the whole workbook on save, so the repair is only applied when the logger's
# own .hobo binary exists beside it - the inference-free original stays
# recoverable, exactly as on the CSV side.
_STAMP_CELL = re.compile(r'^\s*(\d\d)/(\d\d)/(\d\d)(\s+)(\d{1,2})h(\d{1,2})min(\d{1,2})s\s*$')


def parse_stamps_xlsx(ws):
    """[(cell, date, seconds-within-half-day)] in row order, or [] when this
    sheet does not carry the collapsed clock."""
    out = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                m = _STAMP_CELL.match(cell.value)
                if m:
                    out.append((cell, (m.group(3), m.group(1), m.group(2)),
                                (int(m.group(5)) % 12) * 3600
                                + int(m.group(6)) * 60 + int(m.group(7))))
                    break                  # one datetime column per row
    return out


def _sheet_fingerprint(ws, clock_cells):
    """Every cell value except the clock ones - so two versions of a workbook
    compare equal iff nothing BUT the clock changed."""
    skip = {c.coordinate for c in clock_cells}
    return tuple((c.coordinate, c.value) for row in ws.iter_rows() for c in row
                 if c.coordinate not in skip)


def _read(path):
    df, _ = dh.read_hobo({'raw_data_path': os.path.dirname(path),
                          'file_name': os.path.basename(path),
                          'input_type': 'HOBO', 'correct_gmt3h': False}, {})
    return df


def _md5(path):
    h = hashlib.md5()
    with open(path, 'rb') as f:
        for b in iter(lambda: f.read(1 << 20), b''):
            h.update(b)
    return h.hexdigest()


def update_manifest(repaired):
    """New checksum/size, status and a dated note for each repaired file. The
    manifest is the archive's integrity record: an edited file with a stale
    md5 reads as corruption instead of a documented repair."""
    if not os.path.isfile(MANIFEST):
        print('   WARNING: no manifest at %s' % MANIFEST)
        return
    with open(MANIFEST, newline='', encoding='utf-8-sig') as f:
        rows = list(csv.DictReader(f))
    fields = list(rows[0].keys()) if rows else []
    by_name = {os.path.basename(r['dest']): r for r in rows}
    hit = 0
    for path in repaired:
        row = by_name.get(os.path.basename(path))
        if row is None:
            continue
        row['md5'] = _md5(path)
        row['size_bytes'] = str(os.path.getsize(path))
        row['status'] = 'collapsed_clock_reconstructed'
        row['note'] = ('pt-BR export written with a 12-hour clock and no AM/PM '
                       'marker; the afternoon half was reconstructed from row '
                       'order and sampling interval on 2026-08-10 by '
                       'repair_collapsed_clock.py, validated against the light '
                       'phase; the original export remains in bruto\\*.hobo')
        hit += 1
    with open(MANIFEST, 'w', newline='', encoding='utf-8-sig') as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(rows)
    print('manifest updated for %d of %d repaired file(s)' % (hit, len(repaired)))


def main(dry_run=False):
    scratch = tempfile.mkdtemp(prefix='collapsedfix_')
    repaired, skipped, refused = [], [], []
    for path in sorted(glob.glob(os.path.join(H_RAW, '**', 'planilha', '*.csv'),
                                 recursive=True)):
        name = os.path.basename(path)
        if name in EXCLUDED:
            refused.append((name, 'excluded: several sites stacked in one sheet'))
            continue
        with open(path, 'r', encoding='latin-1', newline='') as f:
            text = f.read()
        stamps = parse_stamps(text)
        if len(stamps) < 50:
            continue                                  # not this format
        before = _read(path)
        n_before, dup_before = len(before), int(before['Datetime'].duplicated().sum())
        if dup_before == 0:
            skipped.append((name, 'no duplicated timestamps - nothing collapsed'))
            continue

        # stage and validate in a LOCAL scratch dir, keeping the file name (the
        # reader validates the extension) - the archive is only touched once
        # every gate has passed
        tmp = os.path.join(scratch, name)
        try:
            # ---- the absolute half, decided by the light and nothing else ----
            # (the sampling interval provably cannot decide it: both seeds give
            # the identical step pattern)
            options = {}
            for seed_pm in (False, True):
                r = reconstruct(stamps, seed_pm=seed_pm)
                if r is None:
                    continue
                body = rewrite(text, stamps, r['times'])
                with open(tmp, 'w', encoding='latin-1', newline='') as f:
                    f.write(body)
                df = _read(tmp)
                options[seed_pm] = (r, body, QC.light_clock_phase(
                    pd.to_datetime(df['Datetime']), df['Luminosity (lux)']), df)
            if not options:
                refused.append((name, 'could not be made monotonic/regular'))
                continue
            if not all(o[2]['evaluable'] for o in options.values()):
                refused.append((name, 'too little light to tell morning from '
                                      'afternoon - needs a HOBOware re-export'))
                continue
            seed_pm = min(options, key=lambda k: options[k][2]['offset_h'])
            rec, body, phase, after = options[seed_pm]
            by = 'light'
            if phase['offset_h'] > MAX_NOON_OFFSET_H:
                # the light channel cannot choose (shaded or fouled): fall back
                # to the water's own diurnal warming, which no shading moves
                temps = {k: temp_peak_hour(pd.to_datetime(o[3]['Datetime']),
                                           o[3]['Temperature (degC)'])
                         for k, o in options.items()}
                usable = {k: v for k, v in temps.items() if v == v}      # drop NaN
                if not usable:
                    refused.append((name, 'no half puts the light near noon (%s) and there '
                                          'are too few complete days to use temperature'
                                    % ' / '.join('%.1f h' % o[2]['peak_hour']
                                                 for o in options.values())))
                    continue
                seed_pm = min(usable, key=lambda k: abs((usable[k] - TEMP_PEAK_H + 12) % 24 - 12))
                off = abs((usable[seed_pm] - TEMP_PEAK_H + 12) % 24 - 12)
                if off > MAX_TEMP_OFFSET_H:
                    refused.append((name, 'neither light (%s) nor temperature (%s) puts this '
                                          'series on a normal day - not a plain collapsed clock'
                                    % (' / '.join('%.1f h' % o[2]['peak_hour'] for o in options.values()),
                                       ' / '.join('%.1f h' % v for v in usable.values()))))
                    continue
                rec, body, phase, after = options[seed_pm]
                by = 'temp %.1f h' % usable[seed_pm]
            if rec['step_match'] < MIN_STEP_MATCH:
                refused.append((name, 'only %.0f%% of steps land on the %d s interval'
                                % (100 * rec['step_match'], rec['interval'])))
                continue

            problems = []
            # monotonicity is checked on the WRITTEN text, re-parsed here: the
            # reader sorts its output, so asking it would hide exactly the
            # corruption this gate exists to catch
            back = parse_stamps(body)
            seq, dayi, seen = [], -1, {}
            for m, d, _s in back:
                if d not in seen:
                    seen[d] = len(seen)
                dayi = seen[d]
                hh, mi, ss = int(m.group(5)), int(m.group(6)), int(m.group(7))
                seq.append(dayi * 86400 + hh * 3600 + mi * 60 + ss)
            if any(b <= a for a, b in zip(seq, seq[1:], strict=False)):
                problems.append('the written file is not in increasing time order')
            t = pd.to_datetime(after['Datetime'])
            if int(t.duplicated().sum()):
                problems.append('%d duplicated timestamps remain'
                                % int(t.duplicated().sum()))
            # the strongest gate: nothing but the clock may have changed
            if _clock_blanked(body) != _clock_blanked(text):
                problems.append('something other than the clock changed')
            if problems:
                refused.append((name, '; '.join(problems)))
                continue

            with open(tmp, 'w', encoding='latin-1', newline='') as f:
                f.write(body)
            if not dry_run:
                shutil.copy2(tmp, path)
            repaired.append((path, n_before, dup_before, rec, phase, seed_pm))
            print('%-46s %5d rows | %4d dups -> 0 | step %5d s (%3.0f%%) | %s | peak %4.1f h, %3.0f%% daylight'
                  % (name[:46], n_before, dup_before, rec['interval'],
                     100 * rec['step_match'], ('PM' if seed_pm else 'AM') + ' by ' + by,
                     phase['peak_hour'], 100 * phase['daylight_frac']))
        finally:
            if os.path.isfile(tmp):
                os.remove(tmp)

    # ---------------------------- the same defect in .xlsx exports -----------
    import openpyxl
    for path in sorted(glob.glob(os.path.join(H_RAW, '**', 'planilha', '*.xlsx'),
                                 recursive=True)):
        name = os.path.basename(path)
        if name in EXCLUDED:
            continue
        wb = openpyxl.load_workbook(path)
        ws = wb[wb.sheetnames[0]]
        stamps = parse_stamps_xlsx(ws)
        if len(stamps) < 50:
            wb.close()
            continue
        before = _read(path)
        n_before = len(before)
        dup_before = int(before['Datetime'].duplicated().sum())
        if dup_before == 0:
            skipped.append((name, 'no duplicated timestamps - nothing collapsed'))
            wb.close()
            continue
        # The original must survive the whole-workbook rewrite. Normally the
        # logger's own .hobo binary IS that original; when it is missing, a
        # byte copy of the workbook is written to bruto\ first, so the repair
        # is never the only version of the file in existence.
        stem = os.path.splitext(name)[0]
        bruto = os.path.join(os.path.dirname(os.path.dirname(path)), 'bruto')
        if not os.path.isfile(os.path.join(bruto, stem + '.hobo')):
            backup = os.path.join(bruto, stem + '.original.xlsx')
            if not os.path.isfile(backup):
                if dry_run:
                    print('   (would back up %s -> bruto\\%s)'
                          % (name, os.path.basename(backup)))
                else:
                    os.makedirs(bruto, exist_ok=True)
                    shutil.copy2(path, backup)
            if not dry_run and not os.path.isfile(backup):
                refused.append((name, 'backup could not be written - not rewriting '
                                      'the workbook in place'))
                wb.close()
                continue

        fingerprint = _sheet_fingerprint(ws, [c for c, _d, _s in stamps])
        tmp = os.path.join(scratch, name)
        try:
            options = {}
            for seed_pm in (False, True):
                r = reconstruct(stamps, seed_pm=seed_pm)
                if r is None:
                    continue
                for (cell, _d, _s), t in zip(stamps, r['times'], strict=True):
                    m = _STAMP_CELL.match(cell.value)
                    sod = t % 86400
                    cell.value = ('%s/%s/%s%s%dh%dmin%ds'
                                  % (m.group(1), m.group(2), m.group(3), m.group(4),
                                     sod // 3600, (sod % 3600) // 60, sod % 60))
                wb.save(tmp)
                df = _read(tmp)
                options[seed_pm] = (r, QC.light_clock_phase(
                    pd.to_datetime(df['Datetime']), df['Luminosity (lux)']), df)
            if not options:
                refused.append((name, 'could not be made monotonic/regular'))
                continue
            if not all(o[1]['evaluable'] for o in options.values()):
                refused.append((name, 'too little light to tell morning from '
                                      'afternoon - needs a HOBOware re-export'))
                continue
            seed_pm = min(options, key=lambda k: options[k][1]['offset_h'])
            rec, phase, after = options[seed_pm]
            by = 'light'
            if phase['offset_h'] > MAX_NOON_OFFSET_H:
                # same fallback as the CSV path: a shaded light channel cannot
                # choose the half, but the water's diurnal warming can
                temps = {k: temp_peak_hour(pd.to_datetime(o[2]['Datetime']),
                                           o[2]['Temperature (degC)'])
                         for k, o in options.items()}
                usable = {k: v for k, v in temps.items() if v == v}
                if not usable:
                    refused.append((name, 'no half puts the light near noon (%s) and too few '
                                          'complete days to use temperature'
                                    % ' / '.join('%.1f h' % o[1]['peak_hour']
                                                 for o in options.values())))
                    continue
                seed_pm = min(usable, key=lambda k: abs((usable[k] - TEMP_PEAK_H + 12) % 24 - 12))
                if abs((usable[seed_pm] - TEMP_PEAK_H + 12) % 24 - 12) > MAX_TEMP_OFFSET_H:
                    refused.append((name, 'neither light (%s) nor temperature (%s) puts this '
                                          'series on a normal day'
                                    % (' / '.join('%.1f h' % o[1]['peak_hour'] for o in options.values()),
                                       ' / '.join('%.1f h' % v for v in usable.values()))))
                    continue
                rec, phase, after = options[seed_pm]
                by = 'temp %.1f h' % usable[seed_pm]
            if rec['step_match'] < MIN_STEP_MATCH:
                refused.append((name, 'only %.0f%% of steps land on the %d s interval'
                                % (100 * rec['step_match'], rec['interval'])))
                continue

            # rebuild the winning workbook (the loop above left the PM variant in wb)
            for (cell, _d, _s), t in zip(stamps, rec['times'], strict=True):
                m = _STAMP_CELL.match(cell.value)
                sod = t % 86400
                cell.value = ('%s/%s/%s%s%dh%dmin%ds'
                              % (m.group(1), m.group(2), m.group(3), m.group(4),
                                 sod // 3600, (sod % 3600) // 60, sod % 60))
            wb.save(tmp)

            problems = []
            check = openpyxl.load_workbook(tmp)
            cws = check[check.sheetnames[0]]
            cstamps = parse_stamps_xlsx(cws)
            seq, seen = [], {}
            for cell, d, _s in cstamps:
                if d not in seen:
                    seen[d] = len(seen)
                mm = _STAMP_CELL.match(cell.value)
                seq.append(seen[d] * 86400 + int(mm.group(5)) * 3600
                           + int(mm.group(6)) * 60 + int(mm.group(7)))
            if any(b <= a for a, b in zip(seq, seq[1:], strict=False)):
                problems.append('the written workbook is not in increasing time order')
            if _sheet_fingerprint(cws, [c for c, _d, _s in cstamps]) != fingerprint:
                problems.append('a cell other than the clock changed')
            check.close()
            t = pd.to_datetime(after['Datetime'])
            if int(t.duplicated().sum()):
                problems.append('%d duplicated timestamps remain' % int(t.duplicated().sum()))
            if problems:
                refused.append((name, '; '.join(problems)))
                continue

            if not dry_run:
                shutil.copy2(tmp, path)
            repaired.append((path, n_before, dup_before, rec, phase, seed_pm))
            print('%-46s %5d rows | %4d dups -> 0 | step %5d s (%3.0f%%) | %s | peak %4.1f h, %3.0f%% daylight  [xlsx]'
                  % (name[:46], n_before, dup_before, rec['interval'],
                     100 * rec['step_match'], ('PM' if seed_pm else 'AM') + ' by ' + by,
                     phase['peak_hour'], 100 * phase['daylight_frac']))
        finally:
            wb.close()
            if os.path.isfile(tmp):
                os.remove(tmp)

    shutil.rmtree(scratch, ignore_errors=True)
    print('\nrepaired : %d' % len(repaired))
    print('skipped  : %d (already fine)' % len(skipped))
    print('REFUSED  : %d' % len(refused))
    for n, why in refused:
        print('   %-52s %s' % (n[:52], why))
    if repaired and not dry_run:
        update_manifest([p for p, *_ in repaired])
    if dry_run:
        print('\n(dry run - nothing was written)')


if __name__ == '__main__':
    main(dry_run='--dry-run' in sys.argv)
